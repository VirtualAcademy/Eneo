import datetime
import os, re, sys, django
import openpyxl
import itertools
import collections


# print(os.path.dirname(os.path.abspath("__file__")))

sys.path.append(os.path.dirname(os.path.abspath("__file__"))) #here is root folder(means parent).
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "fmms.settings")
django.setup()
#

from fuel.models import Powerplant
from production.models import Generators, PowerAvailable

fpath = r"D:\Data\Eneo\dev\fmms\media\imports"


class WorkBook(object):

    def __init__(self,filepath, filename):
        self.FILEPATH = filepath
        self.FILENAME = filename
        self.OPENWORKBOOK = self.readFile()
        self.SHEETTITLES = self.getSheetsNames()
        # self.ordeddictionary = collections.OrderedDict
        self.NULLCELLVALUES = lambda row:map(lambda cell:cell.value,row) # maps a list of rows of list of cells in that row, in order to extract its values
        self.NONNULLCELLVALUES = lambda cell_values: itertools.filterfalse(lambda cell_value: cell_value is None, cell_values)

    def readFile(self):
        return openpyxl.load_workbook(os.path.join(self.FILEPATH, self.FILENAME), data_only=True)

    def makeZip(self):
        opened_sheet = self.getSheetByName('Suivi Conso HFO')
        return opened_sheet.iter_cols(max_row=2, max_col=9)

    def getSheetsNames(self):
        return self.OPENWORKBOOK.get_sheet_names()

    def getPlan(self, opened_plansheet):
        for index, time_column in enumerate(opened_plansheet['B6':'B173']):
            print(index, time_column[0].value,*opened_plansheet['E%d'%(index)].value)

    # def sheetNametodate(self):
    #     self.listofdates = []
    #     for i in range(len(self.sheetsname)):
    #         try:
    #             cleanv = self.sheetsname[i].replace('-','').split()[0]
    #             if not(re.match(r'\d+',cleanv)): #filtring the digits from text
    #                 print('none')
    #                 continue
    #             else:
    #                 dd = cleanv[:2]
    #                 mm = cleanv[2:4]
    #                 yyyy = cleanv[4:]
    #                 if len(yyyy)<4:
    #                     yyyy='2018'
    #                 print(([dd,mm,yyyy,cleanv]))
    #                 self.listofdates.append(datetime.date(int(yyyy),int(mm),int(dd)))
    #         except ValueError:
    #             continue
    #     return self.listofdates
        # print('count is {} and sheet is {}'.format(str(n),str(len(self.sheetsname))))

    # def rdsheetvalues(self,sheetname):
    #     readsheet =self.workbk.get_sheet_by_name(sheetname)
    #     # print([readsheet['H1'].value,readsheet['H4'].value,readsheet['J1'].value,readsheet['K1'].value])
    #     v=lambda s:list(map(lambda x:x.value,s)) # using map to iterate over a list of list
    #     for i,r in enumerate(readsheet.rows):
    #         print([i,v(r)])

    def getSheetByName(self,sheetname):
        return self.OPENWORKBOOK.get_sheet_by_name(sheetname)

    def openCollectFile(self,opened_sheet):
        # allrows = rdsheet.rows
        # p_plants=set(self.lambdax(allrows[:1]))
        # type_fuel=set(self.lambdax(list(allrows)[0]))
        for index, row in enumerate(opened_sheet.iter_rows):
            if index == 0:
                all_power_plant_names = list(self.NONNULLCELLVALUES(self.NULLCELLVALUES(row)))
            elif index==1:
                fuel_type_list =list( self.NULLCELLVALUES(row))[:10]
            elif index==2:
                titles = list(self.NULLCELLVALUES(row))
                break
        return titles[1:10], all_power_plant_names
        # for i,r in enumerate(rdsheet.rows):
        #     # if i==
        #     print([i,r])

    def avalablePowerSheet(self, opened_powersheet):
        """
        Returns a dictionary with keys -->
        ['location', 'category', 'plant_name', 'date',
        'values:(Hours, Total Availe Power, Availe Power per generator)']
        """

        # data = opened_powersheet[:4]
        values = [list(self.NULLCELLVALUES(u)) for u in opened_powersheet.iter_rows(min_row=3, max_row=26)]
        date = opened_powersheet['A1'].value
        plantname = opened_powersheet['B1'].value

        categorymatch= re.match(r'.+(HFO|LFO|HYDR).+',opened_powersheet['B1'].value)
        if not(categorymatch):
            print('hasn not got a category')
            category ='Others'
        else:
            category=categorymatch.group(1)
            if not(category == 'HYDR'):
                category = 'THERM'
            else:
                category = category
        location = opened_powersheet.title
        # centrale = opened_powersheet['A3'].value
        # pname = Powerplant.objects.get(location='limbe')
        # max_col = len(list(data))
        # s = list(self.NULLCELLVALUES(data))
        return collections.OrderedDict({'date':date, 'plant_name':plantname, 'category':category,'location':location,'values':values})
        #,centrale, [list(self.NULLCELLVALUES(u)) for u in opened_powersheet.iter_rows(min_row=3, max_row=26)]






        # d = datetime.date(*[int(i) for i in self.sheetsname[0].split('-')[::-1]])
b = WorkBook(fpath,'plan.xlsx')
# print(b.SHEETTITLES)
# for sheet in b.sheetsname:
#     if not(re.match(r'\d+',sheet.split()[0])):
#         continue
#
#     print(b.rdsheetvalues(sheet))
#     try:
#
# #         # cleanv = self.sheetsname[i].replace('-', '').split()[0]
#         if not (re.match(r'\d+',re.sub(r'\D',sheet.split()[0]))):  # filtring the digits from text
#             print(sheet)
#             continue
# #
#         print(b.rdsheetvalues(sheet))
#     except:
#         continue
# print(b.rdsheetvalues('Suivi Conso HFO'))
# print(b.openCollectFile())
# print(list(b.makeZip()))


# for sheet in b.SHEETTITLES:
#     data = b.avalablePowerSheet(b.getSheetByName(sheet))
#     location = data['location']
#     cat = data['category']
#     name = data['plant_name']
#     date = data['date']
#     group_data = data['values']
#     print('For {}'.format(sheet.capitalize()))
#     # print(group_data)
#     for group in group_data:
#         ptime = group[0]
#         total_power = group[1]
#
#         for grpnum, power in enumerate(group[2:]):
#             try:
#                 plant_gotten_or_created = Powerplant.objects.get_or_create(location=location,defaults={'plant_name': name,
#                                                                                                        'production_capacity': 0,
#                                                                                                        'category': cat})
#
#                 power_plant = plant_gotten_or_created[0]
#
#                 available_power = PowerAvailable(location=location, power_plant=power_plant, power_units= total_power,
#                                                  date_recorded=date, time_recorded=ptime)
#                 if not(power):
#                     continue
#
#                 # if power == 'Réserve':
#                 #     power = 0
#                 generator = Generators(location=location, group_number=grpnum+1, power_units=power,
#                                        date_recorded=date, time_recorded=ptime)
#
#             except:
#
#                 print('Not successful in creating {} data, an error occurred'.format(name))
#                 raise
#
#             finally:
#                 power_plant.save()
#                 available_power.save()
#                 generator.save()
#                 print("Group {} and Power Available {} created at {} for {}".format(grpnum+1, total_power, ptime, name))
#
#     print('Done.')
# g=b.avalablePowerSheet(b.getSheetByName('Limbe'))
# print(g['location'])
# print(g['category'])
# print(g['plant_name'])
# print(g['date'])
# print(g['values'])

b.getPlan(b.getSheetByName('Prévisions horaires'))
